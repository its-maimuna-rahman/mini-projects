"""
utils_storage.py  —  BankOS
SQLite + CSV persistence layer.

first_run_setup() returns:
    (conn, acc_log_conn, tx_log_conn, pending_transfer_conn)
so main.py can pass the correct CSV paths to every logging call.

All monetary amounts are in minor units (integer).
"""

from __future__ import annotations

import csv
import sqlite3
from datetime import datetime
from pathlib import Path

from utils_currency import SUPPORTED, convert_minor, format_money
from utils_security import (
    MAX_ATTEMPTS,
    VAULT_MAX_TRIES,
    compute_lockout_until,
    hash_password,
    is_locked,
    now_display,
)
from account_cls import Account, CreditCard, Non_Credit_Card, Vault

# ── constants ─────────────────────────────────────────────────────────────────

# Pending threshold in minor units:
#   1 000.00 USD  → 100_000 minor  (cents)
# 120 000.00 BDT  → 12_000_000 minor  (paise)
PENDING_THRESHOLD: dict[str, int] = {
    "USD": 100_000,
    "BDT": 12_000_000,
}

# Used when re-executing an admin-approved transfer so it can never re-queue.
_APPROVE_BYPASS = 10 ** 15

VALID_CATEGORIES: set[str] = {"food", "bills", "shopping", "transfer", "other"}

# ── filter map (exported for main.py) ────────────────────────────────────────

FILTER_WHERE: dict[str, str] = {
    "all_account":             "1=1",
    "credit_card_account":     "acc_type='credit_card'",
    "non_credit_card_account": "acc_type='non_credit_card'",
    "vault_account":           "vault_no IS NOT NULL",
    "non_vault_account":       "vault_no IS NULL",
    "usd_accounts":            "currency='USD'",
    "bdt_accounts":            "currency='BDT'",
}

# ── DB schema ─────────────────────────────────────────────────────────────────

SCHEMA: list[str] = [
    """CREATE TABLE IF NOT EXISTS accounts (
        acc_num                     INTEGER PRIMARY KEY,
        acc_password_hash           TEXT    NOT NULL,
        acc_password_plain_debug    TEXT,
        acc_balance                 INTEGER NOT NULL DEFAULT 0,
        currency                    TEXT    NOT NULL CHECK(currency IN ('BDT','USD')),
        acc_type                    TEXT    NOT NULL
                                    CHECK(acc_type IN ('credit_card','non_credit_card'))
                                    DEFAULT 'non_credit_card',
        credit_card_num             TEXT,
        credit_card_pin_hash        TEXT,
        credit_card_pin_plain_debug TEXT,
        credit_card_limit           INTEGER DEFAULT 0,
        credit_used                 INTEGER DEFAULT 0,
        vault_no                    TEXT,
        vault_password_hash         TEXT,
        vault_password_plain_debug  TEXT,
        vault_balance               INTEGER DEFAULT 0,
        is_frozen                   INTEGER DEFAULT 0,
        daily_transfer_limit        INTEGER DEFAULT 500000,
        failed_login_attempts       INTEGER DEFAULT 0,
        locked_until                TEXT
    )""",
    """CREATE TABLE IF NOT EXISTS pending_transfers (
        id               INTEGER PRIMARY KEY AUTOINCREMENT,
        timestamp        TEXT    NOT NULL,
        sender_acc_num   INTEGER NOT NULL,
        receiver_acc_num TEXT    NOT NULL,
        amount           INTEGER NOT NULL,
        currency         TEXT    NOT NULL,
        via_credit       INTEGER NOT NULL DEFAULT 0,
        status           TEXT    NOT NULL DEFAULT 'pending',
        reviewed_at      TEXT
    )""",
]


# =============================================================================
#  DB initialisation
# =============================================================================

def init_storage(db_path: str) -> sqlite3.Connection:
    """Open/create SQLite DB, apply schema, return connection with row_factory."""
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    for stmt in SCHEMA:
        conn.execute(stmt)
    conn.commit()
    return conn


# =============================================================================
#  First-run setup wizard
# =============================================================================

def first_run_setup() -> tuple[
    sqlite3.Connection, sqlite3.Connection, sqlite3.Connection,
    sqlite3.Connection, sqlite3.Connection,
]:
    """
    Interactive first-run wizard.
    Returns (conn, acc_log_conn, tx_log_conn, freeze_log_conn, pending_conn).

    Flow (per outline_README.md lines 107-157):
      === BANK MANAGEMENT SYSTEM ===
      [1] Import existing database
      [2] Start fresh database
      [3] Exit
      → Optional CSV/XLSX import into DB

      --- Log Table Setup ---
      Account log setup:         same [1]/[2]/[3] + optional CSV/XLSX import
      Transaction log setup:     same flow
      Account Freeze data setup: same flow
      Pending transfers setup:   same flow
    """
    print("\n=== BANK MANAGEMENT SYSTEM ===")
    print("  [1] Import existing database")
    print("  [2] Start fresh database")
    print("  [3] Exit")

    while True:
        choice = input("  Choose [1/2/3]: ").strip()
        if choice in ("1", "2", "3"):
            break
        print("  [!] Enter 1, 2, or 3.")

    if choice == "3":
        raise SystemExit(0)

    if choice == "1":
        while True:
            db_name = input("  Enter .db filename to import: ").strip()
            if not db_name:
                print("  [!] Filename cannot be empty.")
                continue
            if not db_name.lower().endswith(".db"):
                print(f"  [!] File must be a .db file (got '{db_name}').")
                continue
            from pathlib import Path as _P
            if not _P(db_name).exists():
                print(f"  [!] File '{db_name}' not found. Try again.")
                continue
            break
        conn = init_storage(db_name)
        print(f"  [OK] Opened existing database: {db_name}")
    else:
        db_name = input("  New database name (Enter for 'bank.db'): ").strip() or "bank.db"
        if not db_name.lower().endswith(".db"):
            db_name += ".db"
        conn = init_storage(db_name)
        print(f"  [OK] Created fresh database: {db_name}")

    # ── Optional CSV / XLSX import ────────────────────────────────────────────
    while True:
        ans = input("\n  Import CSV/XLSX files into DB? [yes/no]: ").strip().lower()
        if ans in ("yes", "y"):
            _run_file_import_wizard(conn, restrict_table="accounts")
            break
        if ans in ("no", "n"):
            break
        print("  [!] Enter yes or no.")

    # ── Log DB setup — each log gets its own .db, same flow as main DB ───────
    print("\n--- Log Table Setup ---")
    acc_log_conn = _setup_log_db_menu(
        log_label="Account log",
        default="account_log.db",
        schema=LOG_SCHEMAS["account_log"],
        table="account_log",
    )
    tx_log_conn = _setup_log_db_menu(
        log_label="Transaction log",
        default="transaction_log.db",
        schema=LOG_SCHEMAS["transaction_log"],
        table="transaction_log",
    )
    freeze_log_conn = _setup_log_db_menu(
        log_label="Account Freeze data",
        default="freezing_account.db",
        schema=LOG_SCHEMAS["freezing_account"],
        table="freezing_account",
    )
    pending_conn = _setup_log_db_menu(
        log_label="Pending transfers",
        default="pending_transfers.db",
        schema=LOG_SCHEMAS["pending_transfers"],
        table="pending_transfers",
    )

    return conn, acc_log_conn, tx_log_conn, freeze_log_conn, pending_conn


# ── Per-log DB schemas ────────────────────────────────────────────────────────

LOG_SCHEMAS: dict[str, str] = {
    "account_log": """CREATE TABLE IF NOT EXISTS account_log (
        id        INTEGER PRIMARY KEY AUTOINCREMENT,
        timestamp TEXT    NOT NULL,
        acc_num   INTEGER,
        action    TEXT    NOT NULL,
        details   TEXT    NOT NULL
    )""",
    "transaction_log": """CREATE TABLE IF NOT EXISTS transaction_log (
        id            INTEGER PRIMARY KEY AUTOINCREMENT,
        timestamp     TEXT    NOT NULL,
        acc_num       INTEGER NOT NULL,
        type          TEXT    NOT NULL,
        amount        INTEGER NOT NULL,
        currency      TEXT    NOT NULL,
        category      TEXT    NOT NULL DEFAULT 'other',
        status        TEXT    NOT NULL,
        balance_after INTEGER NOT NULL
    )""",
    "freezing_account": """CREATE TABLE IF NOT EXISTS freezing_account (
        id        INTEGER PRIMARY KEY AUTOINCREMENT,
        timestamp TEXT    NOT NULL,
        acc_num   INTEGER NOT NULL,
        action    TEXT    NOT NULL,
        details   TEXT    NOT NULL
    )""",
    "pending_transfers": """CREATE TABLE IF NOT EXISTS pending_transfers (
        id               INTEGER PRIMARY KEY AUTOINCREMENT,
        timestamp        TEXT    NOT NULL,
        sender_acc_num   INTEGER NOT NULL,
        receiver_acc_num TEXT    NOT NULL,
        amount           INTEGER NOT NULL,
        currency         TEXT    NOT NULL,
        via_credit       INTEGER NOT NULL DEFAULT 0,
        status           TEXT    NOT NULL DEFAULT 'pending',
        reviewed_at      TEXT
    )""",
}


def _init_log_db(db_path: str, schema_sql: str) -> sqlite3.Connection:
    """Open/create a single-table log DB, apply schema, return connection."""
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute(schema_sql)
    conn.commit()
    return conn


def _setup_log_db_menu(
    log_label: str,
    default: str,
    schema: str,
    table: str,
) -> sqlite3.Connection:
    """
    Identical flow to the main DB setup, applied to a single log DB.

      [1] Import existing .db file
            → loop asking filename until the file exists on disk
      [2] Start fresh .db file
            → ask for name (Enter = use default), create it with schema
      [3] Exit
            → exits the whole program

    After the DB is set up, asks:
      "Import CSV/XLSX data into this log? [yes/no]"
      If yes → runs _run_file_import_wizard scoped to this log's table.

    Returns an open sqlite3.Connection for the log DB.
    """
    print(f"\n  {log_label} setup:")
    print(f"    [1] Import existing database")
    print(f"    [2] Start fresh (no import)")
    print(f"    [3] Exit")

    while True:
        sub = input("    Choose [1/2/3]: ").strip()
        if sub in ("1", "2", "3"):
            break
        print("    [!] Enter 1, 2, or 3.")

    if sub == "3":
        raise SystemExit(0)

    if sub == "1":
        while True:
            db_name = input(f"    Enter .db filename to import: ").strip()
            if not db_name:
                print("    [!] Filename cannot be empty.")
                continue
            if not db_name.lower().endswith(".db"):
                print(f"    [!] File must be a .db file (got '{db_name}').")
                continue
            if not Path(db_name).exists():
                print(f"    [!] File '{db_name}' not found. Try again.")
                continue
            break
        log_conn = _init_log_db(db_name, schema)
        print(f"    [OK] Opened existing {log_label.lower()} database: {db_name}")
    else:  # sub == "2"
        db_name = input(f"    New database name (Enter for '{default}'): ").strip() or default
        if not db_name.lower().endswith(".db"):
            db_name += ".db"
        log_conn = _init_log_db(db_name, schema)
        print(f"    [OK] Created fresh {log_label.lower()} database: {db_name}")

    # ── Optional CSV/XLSX import into this log DB ─────────────────────────────
    while True:
        ans = input(f"\n    Import CSV/XLSX files into DB? [yes/no]: ").strip().lower()
        if ans in ("yes", "y"):
            _run_file_import_wizard(log_conn, restrict_table=table)
            break
        if ans in ("no", "n"):
            break
        print("    [!] Enter yes or no.")

    return log_conn


def _run_file_import_wizard(conn: sqlite3.Connection, restrict_table: str | None = None) -> None:
    """
    Import wizard: asks how many files, then for each file:
      1. Ask which format: [1] CSV  [2] XLSX
      2. Ask which table to import into (numbered menu)
         - If restrict_table is set, skip the menu and use that table directly.
      3. Ask for the filename — must exist on disk AND match the chosen format
      4. On failure: ask to retry (count stays frozen) or skip (count advances)
      Duplicates / invalid rows are skipped with a warning.
    """
    IMPORTABLE = ["accounts", "account_log", "transaction_log",
                  "freezing_account", "pending_transfers"]

    # How many files?
    while True:
        raw = input("  How many files to import? : ").strip()
        try:
            count = int(raw)
            if count >= 1:
                break
            print("  [!] Enter a number greater than 0.")
        except ValueError:
            print("  [!] Enter a positive integer.")

    done = 0
    while done < count:
        print(f"\n  --- File {done + 1} of {count} ---")

        # Step 1: choose format
        print("  File format:")
        print("    [1] CSV  (.csv)")
        print("    [2] XLSX (.xlsx)")
        while True:
            fmt = input("    Choose [1/2]: ").strip()
            if fmt == "1":
                ext = ".csv"
                break
            if fmt == "2":
                ext = ".xlsx"
                break
            print("    [!] Enter 1 or 2.")

        # Step 2: choose target table (skip menu if restricted to one table)
        if restrict_table:
            table = restrict_table
            print(f"  Target table: {table}")
        else:
            print("  Target table:")
            for i, t in enumerate(IMPORTABLE, 1):
                print(f"    [{i}] {t}")
            table = ""
            while not table:
                raw = input(f"    Table [1-{len(IMPORTABLE)}]: ").strip()
                try:
                    idx = int(raw) - 1
                    if 0 <= idx < len(IMPORTABLE):
                        table = IMPORTABLE[idx]
                    else:
                        print(f"    [!] Enter a number 1–{len(IMPORTABLE)}.")
                except ValueError:
                    print(f"    [!] Enter a number 1–{len(IMPORTABLE)}.")

        # Step 3: filename — must exist and have correct extension
        while True:
            if ext == ".csv":
                filename = input(f"  Enter .csv file name : ").strip()
            else:
                filename = input(f"  Enter .xlsx file name : ").strip()

            if not filename:
                print("  [!] Filename cannot be empty.")
                continue

            if not filename.lower().endswith(ext):
                print(f"  [!] File must be a {ext} file (got '{filename}').")
                continue

            from pathlib import Path as _P
            if not _P(filename).exists():
                print(f"  [!] File '{filename}' not found on disk.")
                while True:
                    retry = input("  Retry? [yes/no]: ").strip().lower()
                    if retry in ("yes", "y"):
                        break
                    if retry in ("no", "n"):
                        print("  [SKIP] Skipping this file.")
                        done += 1
                        break
                    print("  [!] Enter yes or no.")
                else:
                    continue  # retry: go back to filename prompt
                break  # skipped: exit inner loop

            # File exists — attempt import
            try:
                inserted, skipped = import_file_to_table(conn, filename, table)
                print(f"  [OK] {inserted} row(s) imported into '{table}'.")
                if skipped:
                    print(f"  [!] {skipped} duplicate/invalid row(s) skipped with warning.")
                done += 1
                break
            except Exception as exc:
                print(f"  [!] Import failed: {exc}")
                while True:
                    retry = input("  Retry this file? [yes/no]: ").strip().lower()
                    if retry in ("yes", "y"):
                        break
                    if retry in ("no", "n"):
                        print("  [SKIP] Skipping this file.")
                        done += 1
                        break
                    print("  [!] Enter yes or no.")
                else:
                    continue
                break

    print(f"\n  [SETUP] Import complete. {done} file(s) processed.")


# =============================================================================
#  Logging
# =============================================================================

def log_account_action(
    conn: sqlite3.Connection,
    log_conn: sqlite3.Connection,
    acc_num: int,
    action: str,
    details: str,
) -> None:
    """Write to the account_log DB."""
    ts = now_display()
    log_conn.execute(
        "INSERT INTO account_log(timestamp, acc_num, action, details) VALUES (?,?,?,?)",
        (ts, acc_num, action, details),
    )
    log_conn.commit()


def log_transaction(
    conn: sqlite3.Connection,
    log_conn: sqlite3.Connection,
    acc_num: int,
    tx_type: str,
    amount: int,
    currency: str,
    category: str,
    status: str,
    balance_after: int,
) -> None:
    """Write to the transaction_log DB."""
    ts = now_display()
    log_conn.execute(
        """INSERT INTO transaction_log
           (timestamp, acc_num, type, amount, currency, category, status, balance_after)
           VALUES (?,?,?,?,?,?,?,?)""",
        (ts, acc_num, tx_type, amount, currency, category, status, balance_after),
    )
    log_conn.commit()


def log_freeze_action(
    conn: sqlite3.Connection,
    log_conn: sqlite3.Connection,
    acc_num: int,
    action: str,
    details: str,
) -> None:
    """Write freeze/unfreeze action to the dedicated freezing_account DB."""
    ts = now_display()
    log_conn.execute(
        "INSERT INTO freezing_account(timestamp, acc_num, action, details) VALUES (?,?,?,?)",
        (ts, acc_num, action, details),
    )
    log_conn.commit()


# =============================================================================
#  Authentication
# =============================================================================

def authenticate_customer(
    conn: sqlite3.Connection,
    acc_num: int,
    password: str,
) -> tuple[bool, str]:
    """
    Full customer login gate:
      1. Account exists?
      2. Frozen?
      3. Locked out?
      4. Password correct? → reset counter
         Wrong?            → increment; lock at MAX_ATTEMPTS (2 min)
    Returns (success, message).
    """
    row = conn.execute(
        "SELECT * FROM accounts WHERE acc_num=?", (acc_num,)
    ).fetchone()
    if not row:
        return False, "Account not found."
    if row["is_frozen"]:
        return False, "Account is frozen by admin. Contact the bank."

    locked, secs = is_locked(row["locked_until"])
    if locked:
        m, s = divmod(secs, 60)
        return False, (f"Account locked. Try again in {m}m {s}s."
                       if m else f"Account locked. Try again in {s}s.")

    if row["acc_password_hash"] == hash_password(password):
        conn.execute(
            "UPDATE accounts SET failed_login_attempts=0, locked_until=NULL WHERE acc_num=?",
            (acc_num,),
        )
        conn.commit()
        return True, "Login successful."

    attempts = row["failed_login_attempts"] + 1
    if attempts >= MAX_ATTEMPTS:
        conn.execute(
            "UPDATE accounts SET failed_login_attempts=?, locked_until=? WHERE acc_num=?",
            (attempts, compute_lockout_until(), acc_num),
        )
        conn.commit()
        return False, "Too many failed attempts. Account locked for 2 minutes."

    conn.execute(
        "UPDATE accounts SET failed_login_attempts=? WHERE acc_num=?",
        (attempts, acc_num),
    )
    conn.commit()
    return False, f"Wrong password. {MAX_ATTEMPTS - attempts} attempt(s) left."


# =============================================================================
#  Daily transfer limit
# =============================================================================

def get_outbound_today(conn: sqlite3.Connection, acc_num: int, tx_log_conn: sqlite3.Connection | None = None) -> int:
    """Sum of successful transfer_out amounts for acc_num today (minor units).
    Queries tx_log_conn (transaction log DB) if provided, else falls back to main conn."""
    db = tx_log_conn if tx_log_conn is not None else conn
    today = datetime.now().strftime("%Y-%m-%d")
    row = db.execute(
        """SELECT COALESCE(SUM(amount), 0) AS s FROM transaction_log
           WHERE acc_num=? AND type='transfer_out'
             AND substr(timestamp,1,10)=? AND status='success'""",
        (acc_num, today),
    ).fetchone()
    return int(row["s"])


# =============================================================================
#  Single-account fund operations
# =============================================================================

def add_funds(
    conn: sqlite3.Connection,
    tx_log_conn: sqlite3.Connection,
    acc_num: int,
    amount_minor: int,
    category: str = "other",
) -> str:
    """Deposit amount to account balance. Logs the transaction."""
    row = conn.execute(
        "SELECT * FROM accounts WHERE acc_num=?", (acc_num,)
    ).fetchone()
    if not row:
        return "Account not found."
    if row["is_frozen"]:
        return "Account is frozen."
    conn.execute(
        "UPDATE accounts SET acc_balance=acc_balance+? WHERE acc_num=?",
        (amount_minor, acc_num),
    )
    conn.commit()
    new_bal = conn.execute(
        "SELECT acc_balance FROM accounts WHERE acc_num=?", (acc_num,)
    ).fetchone()["acc_balance"]
    log_transaction(conn, tx_log_conn, acc_num, "add", amount_minor,
                    row["currency"], category, "success", new_bal)
    return (f"Deposited {format_money(amount_minor, row['currency'])}. "
            f"New balance: {format_money(new_bal, row['currency'])}.")


def deduct_funds(
    conn: sqlite3.Connection,
    tx_log_conn: sqlite3.Connection,
    acc_num: int,
    amount_minor: int,
    category: str = "other",
) -> str:
    """Deduct amount from account balance. Logs the transaction."""
    row = conn.execute(
        "SELECT * FROM accounts WHERE acc_num=?", (acc_num,)
    ).fetchone()
    if not row:
        return "Account not found."
    if row["is_frozen"]:
        return "Account is frozen."
    if row["acc_balance"] < amount_minor:
        log_transaction(conn, tx_log_conn, acc_num, "deduct", amount_minor,
                        row["currency"], category, "failed", row["acc_balance"])
        return "Insufficient balance."
    conn.execute(
        "UPDATE accounts SET acc_balance=acc_balance-? WHERE acc_num=?",
        (amount_minor, acc_num),
    )
    conn.commit()
    new_bal = conn.execute(
        "SELECT acc_balance FROM accounts WHERE acc_num=?", (acc_num,)
    ).fetchone()["acc_balance"]
    log_transaction(conn, tx_log_conn, acc_num, "deduct", amount_minor,
                    row["currency"], category, "success", new_bal)
    warning = (f" [WARNING] Low balance: {format_money(new_bal, row['currency'])}."
               if new_bal <= 50_000 else "")
    return f"Deducted {format_money(amount_minor, row['currency'])}.{warning}"


# =============================================================================
#  Transfer engine
# =============================================================================

def _execute_transfer(
    conn: sqlite3.Connection,
    tx_log_conn: sqlite3.Connection,
    sender_num: int,
    receiver_num: int,
    amount_minor: int,
    category: str,
    via_credit: bool,
    threshold_override: int | None = None,
    pending_conn: sqlite3.Connection | None = None,
) -> str:
    """
    Core 1-to-1 transfer logic shared by all transfer modes and review_pending.

    threshold_override:
        None          → use PENDING_THRESHOLD (normal path)
        _APPROVE_BYPASS → admin-approved path; skips daily-limit + threshold checks

    pending_conn:
        Connection to the pending_transfers DB. If None, falls back to conn.
    """
    pdb = pending_conn if pending_conn is not None else conn

    s = conn.execute(
        "SELECT * FROM accounts WHERE acc_num=?", (sender_num,)
    ).fetchone()
    r = conn.execute(
        "SELECT * FROM accounts WHERE acc_num=?", (receiver_num,)
    ).fetchone()

    if not s:
        return "Sender account not found."
    if not r:
        return "Receiver account not found."
    if s["is_frozen"]:
        return "Transfer blocked — sender account is frozen."
    if r["is_frozen"]:
        return "Transfer blocked — receiver account is frozen."

    admin_bypass = (threshold_override == _APPROVE_BYPASS)

    # ── daily limit check (skip for admin-approved) ───────────────────────────
    if not admin_bypass:
        spent_today = get_outbound_today(conn, sender_num, tx_log_conn)
        if spent_today + amount_minor > s["daily_transfer_limit"]:
            remaining = max(0, s["daily_transfer_limit"] - spent_today)
            return (f"Daily limit exceeded. "
                    f"Remaining today: {format_money(remaining, s['currency'])}.")

    # ── pending threshold check (skip for admin-approved) ────────────────────
    threshold = (threshold_override
                 if threshold_override is not None
                 else PENDING_THRESHOLD.get(s["currency"], PENDING_THRESHOLD["BDT"]))
    if amount_minor >= threshold:
        pdb.execute(
            """INSERT INTO pending_transfers
               (timestamp, sender_acc_num, receiver_acc_num, amount,
                currency, via_credit, status)
               VALUES (?,?,?,?,?,?,'pending')""",
            (now_display(), sender_num, str(receiver_num),
             amount_minor, s["currency"], int(via_credit)),
        )
        pdb.commit()
        log_transaction(conn, tx_log_conn, sender_num, "transfer_out", amount_minor,
                        s["currency"], category, "pending", s["acc_balance"])
        return (f"Transfer of {format_money(amount_minor, s['currency'])} exceeds "
                f"threshold. Queued for admin approval.")

    # ── deduct from sender ────────────────────────────────────────────────────
    if via_credit:
        if s["acc_type"] != "credit_card":
            return "Sender has no credit card."
        if s["credit_used"] + amount_minor > s["credit_card_limit"]:
            return "Credit limit insufficient for this transfer."
        conn.execute(
            "UPDATE accounts SET credit_used=credit_used+? WHERE acc_num=?",
            (amount_minor, sender_num),
        )
        sender_bal_after = s["acc_balance"]   # balance unchanged when paying via credit
    else:
        if s["acc_balance"] < amount_minor:
            return "Insufficient balance."
        conn.execute(
            "UPDATE accounts SET acc_balance=acc_balance-? WHERE acc_num=?",
            (amount_minor, sender_num),
        )
        sender_bal_after = s["acc_balance"] - amount_minor

    # ── credit receiver (auto-convert currency if needed) ────────────────────
    recv_amount = convert_minor(amount_minor, s["currency"], r["currency"])
    conn.execute(
        "UPDATE accounts SET acc_balance=acc_balance+? WHERE acc_num=?",
        (recv_amount, receiver_num),
    )
    conn.commit()

    receiver_bal_after = conn.execute(
        "SELECT acc_balance FROM accounts WHERE acc_num=?", (receiver_num,)
    ).fetchone()["acc_balance"]

    log_transaction(conn, tx_log_conn, sender_num, "transfer_out", amount_minor,
                    s["currency"], category, "success", sender_bal_after)
    log_transaction(conn, tx_log_conn, receiver_num, "transfer_in", recv_amount,
                    r["currency"], category, "success", receiver_bal_after)

    msg = f"Transfer successful. {format_money(amount_minor, s['currency'])} sent."
    if s["currency"] != r["currency"]:
        msg += f" Receiver got {format_money(recv_amount, r['currency'])}."
    return msg


def transfer(
    conn: sqlite3.Connection,
    tx_log_conn: sqlite3.Connection,
    sender_num: int,
    receiver_num: int,
    amount_minor: int,
    category: str = "transfer",
    via_credit: bool = False,
    pending_conn: sqlite3.Connection | None = None,
) -> str:
    """Public 1-to-1 transfer with daily limit + pending threshold."""
    return _execute_transfer(conn, tx_log_conn, sender_num, receiver_num,
                             amount_minor, category, via_credit,
                             pending_conn=pending_conn)


def transfer_1tomany(
    conn: sqlite3.Connection,
    tx_log_conn: sqlite3.Connection,
    sender_num: int,
    receiver_nums: list[int],
    amount_minor: int,
    category: str = "transfer",
    via_credit: bool = False,
    pending_conn: sqlite3.Connection | None = None,
) -> list[str]:
    """Transfer amount_minor from one sender to each receiver. Returns one result string per receiver."""
    return [
        f"→ Acc {rec}: "
        f"{_execute_transfer(conn, tx_log_conn, sender_num, rec, amount_minor, category, via_credit, pending_conn=pending_conn)}"
        for rec in receiver_nums
    ]


def transfer_manyto1(
    conn: sqlite3.Connection,
    tx_log_conn: sqlite3.Connection,
    sender_nums: list[int],
    receiver_num: int,
    amount_minor: int,
    category: str = "transfer",
    via_credit: bool = False,
    pending_conn: sqlite3.Connection | None = None,
) -> list[str]:
    """Transfer amount_minor from each sender to one receiver."""
    return [
        f"→ Acc {sen}: {_execute_transfer(conn, tx_log_conn, sen, receiver_num, amount_minor, category, via_credit, pending_conn=pending_conn)}"
        for sen in sender_nums
    ]


# =============================================================================
#  Pending transfer review (admin)
# =============================================================================

def get_pending_transfers(conn: sqlite3.Connection) -> list[sqlite3.Row]:
    """Return all pending_transfers rows with status='pending'."""
    return conn.execute(
        "SELECT * FROM pending_transfers WHERE status='pending' ORDER BY timestamp"
    ).fetchall()


def review_pending(
    conn: sqlite3.Connection,
    tx_log_conn: sqlite3.Connection,
    transfer_id: int,
    approve: bool,
    pending_conn: sqlite3.Connection | None = None,
) -> str:
    """
    Admin approves or rejects a pending transfer.

    Approve:
        Calls _execute_transfer with _APPROVE_BYPASS so the transfer can never
        re-enter the pending queue or be blocked by the daily limit.
        Marks the row 'approved' on success, 'rejected' on failure.

    Reject:
        Reverses the credit_used charge if the transfer was via_credit.
        Marks the row 'rejected'.
    """
    pdb = pending_conn if pending_conn is not None else conn
    row = pdb.execute(
        "SELECT * FROM pending_transfers WHERE id=? AND status='pending'",
        (transfer_id,),
    ).fetchone()
    if not row:
        return "Pending transfer not found or already reviewed."

    ts_now = now_display()

    if not approve:
        # reverse credit charge if applicable
        if row["via_credit"]:
            conn.execute(
                "UPDATE accounts SET credit_used=MAX(0,credit_used-?) WHERE acc_num=?",
                (row["amount"], row["sender_acc_num"]),
            )
        pdb.execute(
            "UPDATE pending_transfers SET status='rejected', reviewed_at=? WHERE id=?",
            (ts_now, transfer_id),
        )
        conn.commit()
        pdb.commit()
        return "Transfer rejected."

    # ── approve: bypass threshold + daily-limit using _APPROVE_BYPASS ─────────
    res = _execute_transfer(
        conn, tx_log_conn,
        int(row["sender_acc_num"]),
        int(row["receiver_acc_num"]),
        int(row["amount"]),
        category="transfer",
        via_credit=bool(row["via_credit"]),
        threshold_override=_APPROVE_BYPASS,
        pending_conn=pending_conn,
    )
    new_status = "approved" if "successful" in res else "rejected"
    pdb.execute(
        "UPDATE pending_transfers SET status=?, reviewed_at=? WHERE id=?",
        (new_status, ts_now, transfer_id),
    )
    pdb.commit()
    return f"[{new_status.upper()}] {res}"


# =============================================================================
#  Vault operations
# =============================================================================

def vault_add(
    conn: sqlite3.Connection,
    tx_log_conn: sqlite3.Connection,
    acc_num: int,
    amount_minor: int,
    vault_password: str,
    from_credit: bool = False,
) -> str:
    """
    Move amount from account balance (or credit card) into the vault.
    Password is verified inside this function (single attempt — caller loops).
    """
    row = conn.execute(
        "SELECT * FROM accounts WHERE acc_num=?", (acc_num,)
    ).fetchone()
    if not row or not row["vault_no"]:
        return "No vault found for this account."
    if row["vault_password_hash"] != hash_password(vault_password):
        return "Wrong vault password."

    if from_credit:
        if row["acc_type"] != "credit_card":
            return "No credit card on this account."
        if row["credit_used"] + amount_minor > row["credit_card_limit"]:
            return "Credit limit insufficient."
        conn.execute(
            "UPDATE accounts SET credit_used=credit_used+?, vault_balance=vault_balance+? "
            "WHERE acc_num=?",
            (amount_minor, amount_minor, acc_num),
        )
    else:
        if row["acc_balance"] < amount_minor:
            return "Insufficient balance."
        conn.execute(
            "UPDATE accounts SET acc_balance=acc_balance-?, vault_balance=vault_balance+? "
            "WHERE acc_num=?",
            (amount_minor, amount_minor, acc_num),
        )

    conn.commit()
    new_bal = conn.execute(
        "SELECT acc_balance FROM accounts WHERE acc_num=?", (acc_num,)
    ).fetchone()["acc_balance"]
    log_transaction(conn, tx_log_conn, acc_num, "vault_add", amount_minor,
                    row["currency"], "other", "success", new_bal)
    return f"Vault deposit: {format_money(amount_minor, row['currency'])}."


def vault_deduct(
    conn: sqlite3.Connection,
    tx_log_conn: sqlite3.Connection,
    acc_num: int,
    amount_minor: int,
    vault_password: str,
    to_credit_payback: bool = False,
) -> str:
    """
    Move amount from vault to account balance, or use it to pay back credit.
    Password is verified inside this function (single attempt — caller loops).
    """
    row = conn.execute(
        "SELECT * FROM accounts WHERE acc_num=?", (acc_num,)
    ).fetchone()
    if not row or not row["vault_no"]:
        return "No vault found for this account."
    if row["vault_password_hash"] != hash_password(vault_password):
        return "Wrong vault password."
    if row["vault_balance"] < amount_minor:
        return "Insufficient vault balance."

    if to_credit_payback:
        if row["acc_type"] != "credit_card":
            return "No credit card on this account."
        paid = min(row["credit_used"], amount_minor)
        conn.execute(
            "UPDATE accounts SET vault_balance=vault_balance-?, credit_used=credit_used-? "
            "WHERE acc_num=?",
            (paid, paid, acc_num),
        )
        conn.commit()
        new_bal = conn.execute(
            "SELECT acc_balance FROM accounts WHERE acc_num=?", (acc_num,)
        ).fetchone()["acc_balance"]
        log_transaction(conn, tx_log_conn, acc_num, "vault_deduct", paid,
                        row["currency"], "other", "success", new_bal)
        return f"Vault → credit payback: {format_money(paid, row['currency'])}."

    conn.execute(
        "UPDATE accounts SET vault_balance=vault_balance-?, acc_balance=acc_balance+? "
        "WHERE acc_num=?",
        (amount_minor, amount_minor, acc_num),
    )
    conn.commit()
    new_bal = conn.execute(
        "SELECT acc_balance FROM accounts WHERE acc_num=?", (acc_num,)
    ).fetchone()["acc_balance"]
    log_transaction(conn, tx_log_conn, acc_num, "vault_deduct", amount_minor,
                    row["currency"], "other", "success", new_bal)
    return f"Vault → balance: {format_money(amount_minor, row['currency'])}."


# =============================================================================
#  Credit card payback
# =============================================================================

def payback_credit(
    conn: sqlite3.Connection,
    tx_log_conn: sqlite3.Connection,
    acc_num: int,
    amount_minor: int,
    from_balance: bool,
    cc_pin: str,
) -> str:
    """
    Pay back credit card debt.
    from_balance=True  → deduct from account balance, reduce credit_used.
    from_balance=False → instant cash (external); only reduce credit_used.
    Requires CC PIN verification.
    """
    row = conn.execute(
        "SELECT * FROM accounts WHERE acc_num=?", (acc_num,)
    ).fetchone()
    if not row or row["acc_type"] != "credit_card":
        return "Credit card account required."
    if row["credit_card_pin_hash"] != hash_password(cc_pin):
        return "Invalid credit card PIN."

    pay = min(amount_minor, row["credit_used"])
    if pay == 0:
        return "No outstanding credit balance."

    if from_balance:
        if row["acc_balance"] < pay:
            return "Insufficient account balance for payback."
        conn.execute(
            "UPDATE accounts SET acc_balance=acc_balance-?, credit_used=credit_used-? "
            "WHERE acc_num=?",
            (pay, pay, acc_num),
        )
        tx_type = "cc_payback_balance"
    else:
        conn.execute(
            "UPDATE accounts SET credit_used=credit_used-? WHERE acc_num=?",
            (pay, acc_num),
        )
        tx_type = "cc_payback_cash"

    conn.commit()
    new_bal = conn.execute(
        "SELECT acc_balance FROM accounts WHERE acc_num=?", (acc_num,)
    ).fetchone()["acc_balance"]
    log_transaction(conn, tx_log_conn, acc_num, tx_type, pay,
                    row["currency"], "other", "success", new_bal)
    return f"Credit payback of {format_money(pay, row['currency'])} successful."


# =============================================================================
#  Account CRUD
# =============================================================================

def create_account(
    conn: sqlite3.Connection,
    acc_num: int,
    password: str,
    currency: str,
    acc_type: str,
    opening_balance: int,
    daily_transfer_limit: int = 500_000,
) -> str:
    """Create a new account. Password is SHA-256 hashed before storage."""
    if currency not in SUPPORTED:
        return f"Unsupported currency '{currency}'."
    if acc_type not in {"credit_card", "non_credit_card"}:
        return "acc_type must be 'credit_card' or 'non_credit_card'."
    if opening_balance < 0:
        return "Opening balance cannot be negative."
    try:
        conn.execute(
            """INSERT INTO accounts
               (acc_num, acc_password_hash, acc_password_plain_debug,
                acc_balance, currency, acc_type, daily_transfer_limit)
               VALUES (?,?,?,?,?,?,?)""",
            (acc_num, hash_password(password), password,
             opening_balance, currency, acc_type, daily_transfer_limit),
        )
        conn.commit()
        return "Account created successfully."
    except sqlite3.IntegrityError:
        return f"Account {acc_num} already exists."


def delete_account(conn: sqlite3.Connection, acc_num: int) -> str:
    """Permanently delete an account by primary key."""
    cur = conn.execute("DELETE FROM accounts WHERE acc_num=?", (acc_num,))
    conn.commit()
    return "Account deleted." if cur.rowcount else "Account not found."


def freeze_account(conn: sqlite3.Connection, acc_num: int, freeze: bool) -> str:
    """Freeze or unfreeze an account (freeze=True to freeze)."""
    row = conn.execute(
        "SELECT acc_num FROM accounts WHERE acc_num=?", (acc_num,)
    ).fetchone()
    if not row:
        return "Account not found."
    conn.execute(
        "UPDATE accounts SET is_frozen=? WHERE acc_num=?", (int(freeze), acc_num)
    )
    conn.commit()
    return f"Account {acc_num} {'frozen' if freeze else 'unfrozen'}."


def set_daily_limit(conn: sqlite3.Connection, acc_num: int, new_limit: int) -> str:
    """Update the daily outbound transfer limit for an account."""
    if new_limit < 0:
        return "Limit cannot be negative."
    row = conn.execute(
        "SELECT currency FROM accounts WHERE acc_num=?", (acc_num,)
    ).fetchone()
    if not row:
        return "Account not found."
    conn.execute(
        "UPDATE accounts SET daily_transfer_limit=? WHERE acc_num=?", (new_limit, acc_num)
    )
    conn.commit()
    return f"Daily transfer limit updated to {format_money(new_limit, row['currency'])}."

def convert_account_type(conn: sqlite3.Connection, acc_num: int, to_type: str) -> str:
    """Convert credit_card ↔ non_credit_card. Assigns default CC fields on upgrade."""
    row = conn.execute(
        "SELECT * FROM accounts WHERE acc_num=?", (acc_num,)
    ).fetchone()
    if not row:
        return "Account not found."
    if to_type == row["acc_type"]:
        return f"Account is already type '{to_type}'."
    if to_type == "credit_card":
        cc_num      = f"4000{acc_num:08d}"
        default_pin = "1234"
        conn.execute(
            """UPDATE accounts SET acc_type='credit_card',
               credit_card_num=?, credit_card_pin_hash=?,
               credit_card_pin_plain_debug=?,
               credit_card_limit=1000000, credit_used=0 WHERE acc_num=?""",
            (cc_num, hash_password(default_pin), default_pin, acc_num),
        )
    else:
        conn.execute(
            """UPDATE accounts SET acc_type='non_credit_card',
               credit_card_num=NULL, credit_card_pin_hash=NULL,
               credit_card_pin_plain_debug=NULL,
               credit_card_limit=0, credit_used=0 WHERE acc_num=?""",
            (acc_num,),
        )
    conn.commit()
    if to_type == "credit_card":
        return (
            f"Account converted to 'credit_card'. "
            f"Card number: {cc_num} | Default PIN: {default_pin} | "
            f"Credit limit: {format_money(1000000, row['currency'])}. "
            f"[IMPORTANT] Tell the customer to change their PIN."
        )
    return f"Account converted to 'non_credit_card'. CC fields cleared."


# =============================================================================
#  Vault CRUD
# =============================================================================

def create_vault(
    conn: sqlite3.Connection,
    acc_num: int,
    vault_no: str,
    vault_password: str,
) -> str:
    """Attach a new vault to an account. vault_no must be globally unique."""
    row = conn.execute(
        "SELECT * FROM accounts WHERE acc_num=?", (acc_num,)
    ).fetchone()
    if not row:
        return "Account not found."
    if row["vault_no"]:
        return "This account already has a vault."
    existing = conn.execute(
        "SELECT acc_num FROM accounts WHERE vault_no=?", (vault_no,)
    ).fetchone()
    if existing:
        return f"Vault number '{vault_no}' is already in use."
    conn.execute(
        """UPDATE accounts SET vault_no=?, vault_password_hash=?,
           vault_password_plain_debug=?, vault_balance=0 WHERE acc_num=?""",
        (vault_no, hash_password(vault_password), vault_password, acc_num),
    )
    conn.commit()
    return f"Vault '{vault_no}' created for account {acc_num}."


def destroy_vault(
    conn: sqlite3.Connection,
    tx_log_conn: sqlite3.Connection,
    acc_num: int,
    vault_password: str,
    transfer_to_balance: bool = True,
) -> str:
    """
    Destroy a vault after verifying the password.
    transfer_to_balance=True  → move remaining funds to account balance.
    transfer_to_balance=False → caller has already handled funds (e.g. CC payback).
    """
    row = conn.execute(
        "SELECT * FROM accounts WHERE acc_num=?", (acc_num,)
    ).fetchone()
    if not row or not row["vault_no"]:
        return "No vault found for this account."
    if row["vault_password_hash"] != hash_password(vault_password):
        return "Wrong vault password. Vault not destroyed."

    vault_balance = row["vault_balance"]
    if vault_balance > 0 and transfer_to_balance:
        conn.execute(
            """UPDATE accounts SET acc_balance=acc_balance+vault_balance,
               vault_no=NULL, vault_password_hash=NULL,
               vault_password_plain_debug=NULL, vault_balance=0 WHERE acc_num=?""",
            (acc_num,),
        )
        new_bal = conn.execute(
            "SELECT acc_balance FROM accounts WHERE acc_num=?", (acc_num,)
        ).fetchone()["acc_balance"]
        log_transaction(conn, tx_log_conn, acc_num, "vault_deduct", vault_balance,
                        row["currency"], "other", "success", new_bal)
    else:
        conn.execute(
            """UPDATE accounts SET vault_no=NULL, vault_password_hash=NULL,
               vault_password_plain_debug=NULL, vault_balance=0 WHERE acc_num=?""",
            (acc_num,),
        )
    conn.commit()
    return f"Vault destroyed. {format_money(vault_balance, row['currency'])} moved to balance."


# =============================================================================
#  OOP loader / saver (for Payment_Processor compatibility if needed)
# =============================================================================

def load_accounts(conn: sqlite3.Connection) -> list[Account]:
    """Reconstruct OOP Account/CreditCard objects with Vault attached."""
    accounts: list[Account] = []
    for row in conn.execute("SELECT * FROM accounts ORDER BY acc_num").fetchall():
        if row["acc_type"] == "credit_card":
            acc = CreditCard(
                acc_num=row["acc_num"],
                password_hash=row["acc_password_hash"],
                acc_balance=row["acc_balance"],
                currency=row["currency"],
                credit_card_num=row["credit_card_num"],
                cc_pin_hash=row["credit_card_pin_hash"],
                credit_card_limit=row["credit_card_limit"],
                credit_used=row["credit_used"],
                is_frozen=bool(row["is_frozen"]),
                daily_transfer_limit=row["daily_transfer_limit"],
            )
        else:
            acc = Non_Credit_Card(
                acc_num=row["acc_num"],
                password_hash=row["acc_password_hash"],
                acc_balance=row["acc_balance"],
                currency=row["currency"],
                is_frozen=bool(row["is_frozen"]),
                daily_transfer_limit=row["daily_transfer_limit"],
            )
        if row["vault_no"]:
            acc.vault = Vault(
                vault_no=row["vault_no"],
                password_hash=row["vault_password_hash"],
                balance=row["vault_balance"],
            )
        accounts.append(acc)
    return accounts


def save_account(conn: sqlite3.Connection, acc: Account) -> None:
    """Upsert a single Account/CreditCard object back to the DB."""
    is_cc = isinstance(acc, CreditCard)
    vault = getattr(acc, "vault", None)
    conn.execute(
        """INSERT OR REPLACE INTO accounts (
            acc_num, acc_password_hash, acc_balance, currency, acc_type,
            credit_card_num, credit_card_pin_hash, credit_card_limit, credit_used,
            vault_no, vault_password_hash, vault_balance,
            is_frozen, daily_transfer_limit
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (
            acc.acc_num, acc._password_hash, acc.acc_balance, acc.currency,
            "credit_card" if is_cc else "non_credit_card",
            acc.credit_card_num   if is_cc else None,
            acc._cc_pin_hash      if is_cc else None,
            acc.credit_card_limit if is_cc else 0,
            acc.credit_used       if is_cc else 0,
            vault.vault_no        if vault else None,
            vault.lock._hash      if vault else None,
            vault.balance         if vault else 0,
            int(acc.is_frozen),
            acc.daily_transfer_limit,
        ),
    )
    conn.commit()


# =============================================================================
#  Listing / filtering / overview
# =============================================================================

def list_accounts(
    conn: sqlite3.Connection,
    filter_key: str = "all_account",
) -> list[sqlite3.Row]:
    """Return account rows matching the given filter key."""
    where = FILTER_WHERE.get(filter_key, "1=1")
    return conn.execute(
        f"SELECT * FROM accounts WHERE {where} ORDER BY acc_num"
    ).fetchall()


def bank_overview(
    conn: sqlite3.Connection,
    pending_conn: sqlite3.Connection | None = None,
) -> dict:
    """
    Return a summary dict:
        totals          → {BDT: int, USD: int}  (minor units)
        frozen_accounts → int
        pending_count   → int
    """
    pdb = pending_conn if pending_conn is not None else conn
    totals = {
        r["currency"]: r["s"]
        for r in conn.execute(
            "SELECT currency, COALESCE(SUM(acc_balance),0) AS s "
            "FROM accounts GROUP BY currency"
        ).fetchall()
    }
    frozen  = conn.execute(
        "SELECT COUNT(*) AS c FROM accounts WHERE is_frozen=1"
    ).fetchone()["c"]
    pending = pdb.execute(
        "SELECT COUNT(*) AS c FROM pending_transfers WHERE status='pending'"
    ).fetchone()["c"]
    return {"totals": totals, "frozen_accounts": frozen, "pending_count": pending}


def get_logs(
    conn: sqlite3.Connection,
    table: str,
    acc_num: int | None = None,
    limit: int = 50,
    acc_log_conn: sqlite3.Connection | None = None,
    tx_log_conn: sqlite3.Connection | None = None,
    freeze_log_conn: sqlite3.Connection | None = None,
    pending_conn: sqlite3.Connection | None = None,
) -> list[sqlite3.Row]:
    """Retrieve rows from account_log, transaction_log, freezing_account, or pending_transfers.
    Uses the dedicated log DB connection if provided, else falls back to main conn."""
    if table not in {"account_log", "transaction_log", "freezing_account", "pending_transfers"}:
        return []
    # Route to the appropriate log DB
    if table == "account_log":
        db = acc_log_conn if acc_log_conn is not None else conn
    elif table == "transaction_log":
        db = tx_log_conn if tx_log_conn is not None else conn
    elif table == "freezing_account":
        db = freeze_log_conn if freeze_log_conn is not None else conn
    else:  # pending_transfers
        db = pending_conn if pending_conn is not None else conn
    # pending_transfers uses sender_acc_num, not acc_num
    if acc_num is None:
        return db.execute(
            f"SELECT * FROM {table} ORDER BY timestamp DESC LIMIT ?", (limit,)
        ).fetchall()
    if table == "pending_transfers":
        return db.execute(
            f"SELECT * FROM {table} WHERE sender_acc_num=? ORDER BY timestamp DESC LIMIT ?",
            (acc_num, limit),
        ).fetchall()
    return db.execute(
        f"SELECT * FROM {table} WHERE acc_num=? ORDER BY timestamp DESC LIMIT ?",
        (acc_num, limit),
    ).fetchall()


def recent_transactions(
    conn: sqlite3.Connection,
    acc_num: int,
    limit: int = 10,
    tx_log_conn: sqlite3.Connection | None = None,
) -> list[sqlite3.Row]:
    """Return the most recent transaction_log rows for an account."""
    db = tx_log_conn if tx_log_conn is not None else conn
    return db.execute(
        "SELECT * FROM transaction_log WHERE acc_num=? ORDER BY timestamp DESC LIMIT ?",
        (acc_num, limit),
    ).fetchall()


# =============================================================================
#  Export (on-demand only — never kept live)
# =============================================================================

def export_accounts_xlsx(
    conn: sqlite3.Connection,
    path: str,
    filter_key: str = "all_account",
) -> str:
    """Export filtered accounts to XLSX. Requires openpyxl."""
    try:
        from openpyxl import Workbook
    except ImportError:
        return "[ERROR] openpyxl not installed. Run: pip install openpyxl"
    rows = list_accounts(conn, filter_key)
    wb = Workbook()
    ws = wb.active
    ws.title = "accounts"
    if rows:
        ws.append(list(rows[0].keys()))
        for r in rows:
            ws.append([r[h] for h in rows[0].keys()])
    wb.save(path)
    return f"[OK] {len(rows)} account(s) exported → {path}"


def export_accounts_csv(
    conn: sqlite3.Connection,
    path: str,
    filter_key: str = "all_account",
) -> str:
    """Export filtered accounts to CSV."""
    rows = list_accounts(conn, filter_key)
    if not rows:
        Path(path).write_text("", encoding="utf-8")
        return f"[OK] No accounts matched '{filter_key}'. Empty file → {path}"
    headers = list(rows[0].keys())
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for r in rows:
            w.writerow([r[h] for h in headers])
    return f"[OK] {len(rows)} account(s) exported → {path}"


# =============================================================================
#  CSV / XLSX import
# =============================================================================

def import_file_to_table(
    conn: sqlite3.Connection,
    file_path: str,
    table: str,
) -> tuple[int, int]:
    """
    Import rows from a CSV or XLSX file into the specified table.
    Returns (inserted, skipped).
    Duplicates and invalid rows are skipped silently (counted in skipped).
    """
    path   = Path(file_path)
    suffix = path.suffix.lower()
    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    if suffix == ".csv":
        with path.open("r", newline="", encoding="utf-8") as f:
            rows = list(csv.DictReader(f))
    elif suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("openpyxl not installed. Run: pip install openpyxl")
        wb   = load_workbook(path)
        ws   = wb.active
        vals = list(ws.values)
        if not vals:
            return 0, 0
        headers = [str(h) for h in vals[0]]
        rows    = [
            dict(zip(headers, r))
            for r in vals[1:]
            if any(v is not None for v in r)
        ]
    else:
        raise ValueError(f"Only .csv and .xlsx supported (got '{suffix}').")

    inserted = skipped = 0
    for row in rows:
        try:
            _insert_row(conn, table, row)
            inserted += 1
        except Exception:
            skipped += 1
    conn.commit()
    return inserted, skipped


def _insert_row(conn: sqlite3.Connection, table: str, row: dict) -> None:
    """Insert one row into the target table. Raises on error."""
    if table == "accounts":
        conn.execute(
            """INSERT OR IGNORE INTO accounts (
                acc_num, acc_password_hash, acc_password_plain_debug,
                acc_balance, currency, acc_type,
                credit_card_num, credit_card_pin_hash, credit_card_pin_plain_debug,
                credit_card_limit, credit_used,
                vault_no, vault_password_hash, vault_password_plain_debug, vault_balance,
                is_frozen, daily_transfer_limit, failed_login_attempts, locked_until
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                int(row["acc_num"]),
                row["acc_password_hash"],
                row.get("acc_password_plain_debug") or None,
                int(row.get("acc_balance", 0)),
                row.get("currency", "BDT"),
                row.get("acc_type", "non_credit_card"),
                row.get("credit_card_num") or None,
                row.get("credit_card_pin_hash") or None,
                row.get("credit_card_pin_plain_debug") or None,
                int(row.get("credit_card_limit", 0)),
                int(row.get("credit_used", 0)),
                row.get("vault_no") or None,
                row.get("vault_password_hash") or None,
                row.get("vault_password_plain_debug") or None,
                int(row.get("vault_balance", 0)),
                int(row.get("is_frozen", 0)),
                int(row.get("daily_transfer_limit", 500_000)),
                int(row.get("failed_login_attempts", 0)),
                row.get("locked_until") or None,
            ),
        )
    elif table == "account_log":
        conn.execute(
            "INSERT INTO account_log(timestamp, acc_num, action, details) VALUES (?,?,?,?)",
            (row["timestamp"], int(row["acc_num"]), row["action"], row["details"]),
        )
    elif table == "transaction_log":
        conn.execute(
            """INSERT INTO transaction_log
               (timestamp, acc_num, type, amount, currency, category, status, balance_after)
               VALUES (?,?,?,?,?,?,?,?)""",
            (
                row["timestamp"], int(row["acc_num"]), row["type"],
                int(row["amount"]), row["currency"],
                row.get("category", "other"),
                row.get("status", "success"),
                int(row["balance_after"]),
            ),
        )
    elif table == "pending_transfers":
        conn.execute(
            """INSERT INTO pending_transfers
               (timestamp, sender_acc_num, receiver_acc_num, amount,
                currency, via_credit, status, reviewed_at)
               VALUES (?,?,?,?,?,?,?,?)""",
            (
                row["timestamp"],
                int(row["sender_acc_num"]),
                str(row["receiver_acc_num"]),
                int(row["amount"]),
                row["currency"],
                int(row.get("via_credit", 0)),
                row.get("status", "pending"),
                row.get("reviewed_at") or None,
            ),
        )
    elif table == "freezing_account":
        conn.execute(
            "INSERT INTO freezing_account(timestamp, acc_num, action, details) VALUES (?,?,?,?)",
            (row["timestamp"], int(row["acc_num"]), row["action"], row["details"]),
        )
    else:
        raise ValueError(f"Unsupported import table: '{table}'")


# =============================================================================
#  Sample data seeder (dev / testing only)
# =============================================================================

def seed_sample_data(conn: sqlite3.Connection) -> None:
    """Insert two sample accounts if the accounts table is empty. No-op otherwise."""
    if conn.execute("SELECT COUNT(*) AS c FROM accounts").fetchone()["c"]:
        return
    conn.execute(
        """INSERT INTO accounts
           (acc_num, acc_password_hash, acc_password_plain_debug,
            acc_balance, currency, acc_type, daily_transfer_limit)
           VALUES (?,?,?,?,?,?,?)""",
        (1001, hash_password("pass1001"), "pass1001",
         2_500_000, "BDT", "non_credit_card", 700_000),
    )
    conn.execute(
        """INSERT INTO accounts
           (acc_num, acc_password_hash, acc_password_plain_debug,
            acc_balance, currency, acc_type,
            credit_card_num, credit_card_pin_hash, credit_card_pin_plain_debug,
            credit_card_limit, credit_used, daily_transfer_limit,
            vault_no, vault_password_hash, vault_password_plain_debug, vault_balance)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (
            2001, hash_password("pass2001"), "pass2001",
            50_000, "USD", "credit_card",
            "4111111122223333", hash_password("7777"), "7777",
            200_000, 0, 50_000,
            "V001", hash_password("vault2001"), "vault2001", 0,
        ),
    )
    conn.commit()
    print("[SEED] Accounts 1001 (BDT) and 2001 (USD+CC+Vault) inserted.")

def transfer_manyto1(
    conn: sqlite3.Connection,
    tx_log_conn: sqlite3.Connection,
    sender_nums: list[int],
    receiver_num: int,
    amount_minor: int,
    category: str = "transfer",
    via_credit: bool = False,
    pending_conn: sqlite3.Connection | None = None,
) -> list[str]:
    """Minimal working many-to-1 transfer (fixes the truncation)."""
    results = []
    for sender_num in sender_nums:
        result = transfer(conn, tx_log_conn, sender_num, receiver_num, amount_minor, category, via_credit, pending_conn=pending_conn)
        results.append(f"From {sender_num}: {result}")
    return results